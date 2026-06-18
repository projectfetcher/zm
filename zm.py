#!/usr/bin/env python3
"""
┌─────────────────────────────────────────────────────────────────┐
│            ZAMBIA JOBS SCRAPER  v2.0  — MimusJobs               │
│  Sources: GoZambiaJobs | ZambiaJob | CVPeopleAfrica |           │
│           JobSearchZM  | GreatZambiaJobs                        │
│  Output:  zambia_jobs.xlsx  (23 fields)                         │
│  Mode:    Pure HTML/Playwright — no external AI APIs            │
└─────────────────────────────────────────────────────────────────┘

USAGE
─────
  pip install -r requirements.txt
  playwright install chromium

  python3 zambia_jobs_scraper_v2.py
  python3 zambia_jobs_scraper_v2.py --max-pages 10
  python3 zambia_jobs_scraper_v2.py --post-wp

ENV VARS (for WordPress posting)
─────────────────────────────────
  WP_URL   = https://zambia.mimusjobs.com
  WP_USER  = admin
  WP_PASS  = your_application_password
"""

import os, re, time, random, logging, argparse, requests
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────
OUTPUT_FILE  = "zambia_jobs.xlsx"
MAX_PAGES    = 5
PAGE_DELAY   = (2, 4)
DETAIL_DELAY = (1, 3)
HEADLESS     = True
BROWSER_ARGS = ["--no-sandbox", "--disable-dev-shm-usage"]

COLUMNS = [
    "Job Title", "Job Type", "Job Qualifications", "Job Experience",
    "Job Location", "Job Field", "Date Posted", "Deadline",
    "Job Description", "Application", "Company URL", "Company Name",
    "Company Logo", "Company Industry", "Company Founded", "Company Type",
    "Company Website", "Company Address", "Company Details",
    "Job URL", "Estimated Deadline", "Salary Range", "Source Site",
]

# ─────────────────────────────────────────────────────────────────
# LOGGING — verbose with field-level output
# ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ZambiaJobs")

# Suppress noisy third-party loggers
for noisy in ("urllib3", "httpcore", "httpx", "hpack", "asyncio"):
    logging.getLogger(noisy).setLevel(logging.WARNING)


def log_job(job):
    """Print every extracted field value for a job — full verbose."""
    bar = "─" * 60
    log.debug(bar)
    for col in COLUMNS:
        val = job.get(col, "")
        if col == "Job Description" and len(val) > 120:
            val = val[:120] + "…"
        status = "✓" if val else "✗"
        log.debug(f"  [{status}] {col:<22}: {val}")
    log.debug(bar)


# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────
def empty_job():
    return {c: "" for c in COLUMNS}


def clean(t):
    return re.sub(r"\s+", " ", str(t or "").strip())


def est_deadline(date_str, days=30):
    try:
        dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
    except Exception:
        dt = datetime.today()
    return (dt + timedelta(days=days)).strftime("%Y-%m-%d")


def normalise_date(raw):
    raw = clean(raw)
    for fmt in (
        "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y",
        "%B %d, %Y", "%d %B %Y", "%b %d, %Y", "%d %b %Y",
    ):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    m = re.search(r"(\d+)\s+day", raw, re.I)
    if m:
        return (datetime.today() - timedelta(days=int(m.group(1)))).strftime("%Y-%m-%d")
    if re.search(r"\d+\s+hour|\btoday\b|\bjust now\b", raw, re.I):
        return datetime.today().strftime("%Y-%m-%d")
    return datetime.today().strftime("%Y-%m-%d")


def extract_experience(text):
    m = re.search(
        r"(\d+[\+]?\s*(?:–|-|to)?\s*\d*\s*years?\s*(?:of\s*)?(?:relevant\s*)?(?:work\s*)?experience)",
        text, re.I
    )
    return clean(m.group(1)) if m else ""


def extract_qualifications(text):
    m = re.search(
        r"((?:Bachelor(?:'s)?|Master(?:'s)?|Diploma|Certificate|PhD|Degree|HND|Grade\s*12)"
        r"[^\n.]{0,120})",
        text, re.I
    )
    return clean(m.group(1))[:150] if m else ""


def extract_salary(text):
    m = re.search(
        r"(ZMW|USD|USD|\$|K)\s?[\d,]+[\s\-–]*(ZMW|USD|\$|K)?\s?[\d,]*"
        r"\s*(?:/\s*(?:month|year|annum|hour|mo|yr))?",
        text, re.I
    )
    return clean(m.group(0))[:80] if m else ""


def extract_deadline(text):
    for pat in [
        r"[Cc]losing\s+[Dd]ate[:\s]+([A-Za-z0-9 ,./\-]+)",
        r"[Dd]eadline[:\s]+([A-Za-z0-9 ,./\-]+)",
        r"[Aa]pply\s+by[:\s]+([A-Za-z0-9 ,./\-]+)",
        r"[Cc]lose[sd]?\s+on[:\s]+([A-Za-z0-9 ,./\-]+)",
    ]:
        m = re.search(pat, text)
        if m:
            return normalise_date(m.group(1).strip()[:30])
    return ""


# ─────────────────────────────────────────────────────────────────
# PLAYWRIGHT BROWSER WRAPPER
# ─────────────────────────────────────────────────────────────────
class Browser:
    def __init__(self):
        self._pw = self._browser = self._ctx = None

    def __enter__(self):
        log.info("🌐 Launching Chromium browser (headless)…")
        self._pw = sync_playwright().start()
        self._browser = self._pw.chromium.launch(
            headless=HEADLESS, args=BROWSER_ARGS
        )
        self._ctx = self._browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-GB",
            extra_http_headers={
                "Accept-Language": "en-GB,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Referer": "https://www.google.com/",
            },
        )
        log.info("✓ Browser ready")
        return self

    def get(self, url, wait="networkidle"):
        page = self._ctx.new_page()
        try:
            log.debug(f"  → GET {url}")
            page.goto(url, wait_until=wait, timeout=35_000)
            time.sleep(random.uniform(*PAGE_DELAY))
            status = page.evaluate("() => document.readyState")
            log.debug(f"  ← readyState={status}  len={len(page.content())}")
            return page
        except PWTimeout:
            log.warning(f"  ⚠ Timeout loading: {url}")
            return page
        except Exception as e:
            log.error(f"  ✗ Navigation error [{url}]: {e}")
            try:
                page.close()
            except Exception:
                pass
            return None

    def __exit__(self, *_):
        try:
            self._browser.close()
            self._pw.stop()
            log.info("🔒 Browser closed")
        except Exception:
            pass


def get_html(page):
    if not page:
        return ""
    try:
        html = page.content()
        return html
    finally:
        try:
            page.close()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────
# SCRAPER 1 — GoZambiaJobs.com
# ─────────────────────────────────────────────────────────────────
def scrape_gozambiajobs(browser):
    SOURCE = "GoZambiaJobs.com"
    BASE   = "https://gozambiajobs.com"
    log.info(f"\n{'═'*60}")
    log.info(f"  SCRAPER 1 — {SOURCE}")
    log.info(f"{'═'*60}")
    jobs, seen_urls = [], set()

    for pg in range(1, MAX_PAGES + 1):
        url = f"{BASE}/jobs?page={pg}"
        log.info(f"\n  📄 Listing page {pg}/{MAX_PAGES}: {url}")
        page = browser.get(url)
        html = get_html(page)
        if not html:
            log.warning("  No HTML returned — stopping pagination")
            break

        sp = BeautifulSoup(html, "lxml")
        job_links = []
        for a in sp.select("a[href]"):
            href = a["href"]
            if re.search(r"/jobs/\d+", href):
                full = href if href.startswith("http") else BASE + href
                if full not in seen_urls:
                    seen_urls.add(full)
                    job_links.append(full)

        if not job_links:
            log.info("  No new job links found — stopping pagination")
            break

        log.info(f"  Found {len(job_links)} job link(s) on page {pg}")
        for i, jurl in enumerate(job_links, 1):
            log.info(f"\n  ┌ Job {i}/{len(job_links)}: {jurl}")
            job = _gozambia_detail(browser, jurl, BASE)
            if job:
                log_job(job)
                jobs.append(job)
            time.sleep(random.uniform(*DETAIL_DELAY))

    log.info(f"\n  ✅ {SOURCE} — total jobs scraped: {len(jobs)}")
    return jobs


def _gozambia_detail(browser, url, base):
    page = browser.get(url)
    html = get_html(page)
    if not html:
        log.warning(f"  No HTML for {url}")
        return None

    sp   = BeautifulSoup(html, "lxml")
    text = sp.get_text(" ", strip=True)
    job  = empty_job()
    job["Job URL"]     = url
    job["Source Site"] = "GoZambiaJobs.com"

    # Title
    h1 = sp.find("h1")
    job["Job Title"] = clean(h1.get_text()) if h1 else ""
    log.debug(f"  Title       → {job['Job Title']}")

    # Company
    ca = sp.select_one("a[href*='/companies/']")
    if ca:
        job["Company Name"] = clean(ca.get_text())
        href = ca["href"]
        job["Company URL"] = href if href.startswith("http") else base + href
    log.debug(f"  Company     → {job['Company Name']}")

    # Logo
    logo = sp.select_one("img[src*='employer'], img[src*='logo']")
    if logo:
        job["Company Logo"] = logo.get("src", "")
    log.debug(f"  Logo        → {job['Company Logo'][:60] if job['Company Logo'] else ''}")

    # Job Type
    for sel in [
        "a[href*='/jobs/full-time']", "a[href*='/jobs/part-time']",
        "a[href*='/jobs/contract']", "a[href*='/jobs/consultancy']",
        "a[href*='/jobs/internship']", "a[href*='/jobs/temporary']",
    ]:
        el = sp.select_one(sel)
        if el:
            job["Job Type"] = clean(el.get_text())
            break
    log.debug(f"  Type        → {job['Job Type']}")

    # Location
    loc = sp.select_one("a[href*='/jobs/in-']")
    if loc:
        job["Job Location"] = clean(loc.get_text())
    log.debug(f"  Location    → {job['Job Location']}")

    # Job Field (category tags)
    tags = sp.select(
        "a[href*='/jobs/'][class*='badge'], "
        "a[href*='/jobs/'][class*='tag'], "
        "span[class*='category'], .job-categories a"
    )
    if tags:
        job["Job Field"] = ", ".join(
            dict.fromkeys(clean(t.get_text()) for t in tags if clean(t.get_text()))
        )
    log.debug(f"  Field       → {job['Job Field']}")

    # Description
    for sel in [
        "div.job-description", "div[class*='description']",
        "div[class*='content']", "section.description", "article",
    ]:
        desc = sp.select_one(sel)
        if desc and len(desc.get_text()) > 100:
            job["Job Description"] = clean(desc.get_text(" "))[:4000]
            break
    log.debug(f"  Desc chars  → {len(job['Job Description'])}")

    # Salary
    job["Salary Range"] = extract_salary(text)
    log.debug(f"  Salary      → {job['Salary Range']}")

    # Date posted
    t = sp.find("time")
    if t:
        job["Date Posted"] = normalise_date(t.get("datetime") or t.get_text())
    log.debug(f"  Date Posted → {job['Date Posted']}")

    # Deadline
    job["Deadline"] = extract_deadline(text)
    log.debug(f"  Deadline    → {job['Deadline']}")

    # Experience & Qualifications
    job["Job Experience"]     = extract_experience(job["Job Description"])
    job["Job Qualifications"] = extract_qualifications(job["Job Description"])
    log.debug(f"  Experience  → {job['Job Experience']}")
    log.debug(f"  Quals       → {job['Job Qualifications'][:60] if job['Job Qualifications'] else ''}")

    # Apply
    apply_a = sp.select_one(
        "a[href*='/apply'], a[class*='apply'], a[id*='apply'], "
        "a.btn-primary[href], a.apply-button"
    )
    if apply_a:
        href = apply_a["href"]
        job["Application"] = href if href.startswith("http") else base + href
    log.debug(f"  Application → {job['Application']}")

    job["Estimated Deadline"] = est_deadline(job["Date Posted"])
    log.debug(f"  Est.Deadlin → {job['Estimated Deadline']}")
    return job


# ─────────────────────────────────────────────────────────────────
# SCRAPER 2 — ZambiaJob.com
# ─────────────────────────────────────────────────────────────────
def scrape_zambiajob(browser):
    SOURCE = "ZambiaJob.com"
    BASE   = "https://www.zambiajob.com"
    log.info(f"\n{'═'*60}")
    log.info(f"  SCRAPER 2 — {SOURCE}")
    log.info(f"{'═'*60}")
    jobs, seen = [], set()

    for pg in range(1, MAX_PAGES + 1):
        url = f"{BASE}/job-vacancies-search-zambia" + (f"?page={pg-1}" if pg > 1 else "")
        log.info(f"\n  📄 Listing page {pg}/{MAX_PAGES}: {url}")
        page = browser.get(url)
        html = get_html(page)
        if not html:
            log.warning("  No HTML — stopping")
            break

        sp = BeautifulSoup(html, "lxml")
        links = [
            (a["href"] if a["href"].startswith("http") else BASE + a["href"])
            for a in sp.select("a[href*='/job-vacancies-zambia/']")
        ]
        new_links = [l for l in links if l not in seen]
        seen.update(new_links)

        if not new_links:
            log.info("  No new job links — stopping")
            break

        log.info(f"  Found {len(new_links)} job link(s)")
        for i, jurl in enumerate(new_links, 1):
            log.info(f"\n  ┌ Job {i}/{len(new_links)}: {jurl}")
            job = _zambiajob_detail(browser, jurl, BASE)
            if job:
                log_job(job)
                jobs.append(job)
            time.sleep(random.uniform(*DETAIL_DELAY))

    log.info(f"\n  ✅ {SOURCE} — total jobs scraped: {len(jobs)}")
    return jobs


def _zambiajob_detail(browser, url, base):
    page = browser.get(url)
    html = get_html(page)
    if not html:
        return None

    sp   = BeautifulSoup(html, "lxml")
    text = sp.get_text(" ", strip=True)
    job  = empty_job()
    job["Job URL"]     = url
    job["Source Site"] = "ZambiaJob.com"

    h1 = sp.find("h1")
    job["Job Title"] = clean(h1.get_text()) if h1 else ""
    log.debug(f"  Title       → {job['Job Title']}")

    field_map = {
        "Company Name": ".company-name, a[href*='/recruiter/'], h2.company, .employer-name",
        "Job Location": ".location, span[class*='region'], .job-location, .city",
        "Job Type":     ".contract-type, .job-type, span[class*='contract'], .employment-type",
        "Job Field":    ".category, a[href*='metiers'], .job-category, .sector",
        "Salary Range": ".salary, .remuneration, [class*='salary']",
    }
    for field, sel in field_map.items():
        el = sp.select_one(sel)
        if el:
            job[field] = clean(el.get_text())
        log.debug(f"  {field:<18}→ {job[field][:70]}")

    # Company URL
    ca = sp.select_one("a[href*='/recruiter/']")
    if ca:
        href = ca["href"]
        job["Company URL"] = href if href.startswith("http") else base + href

    # Logo
    logo = sp.select_one(
        "img.company-logo, .logo img, img[alt*='logo'], .employer-logo img"
    )
    if logo:
        job["Company Logo"] = logo.get("src", "")
    log.debug(f"  Logo        → {job['Company Logo'][:60] if job['Company Logo'] else ''}")

    # Date
    date_el = sp.select_one(".date, time, .published, span[class*='date'], .post-date")
    raw_date = (date_el.get("datetime") or date_el.get_text()) if date_el else ""
    if not raw_date:
        m = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
        raw_date = m.group(1) if m else ""
    job["Date Posted"] = normalise_date(raw_date)
    log.debug(f"  Date Posted → {job['Date Posted']}")

    # Deadline
    job["Deadline"] = extract_deadline(text)
    log.debug(f"  Deadline    → {job['Deadline']}")

    # Description
    for sel in [
        ".offer-details", ".job-description", "article",
        ".content-offer", "div[class*='description']", ".job-content",
    ]:
        desc = sp.select_one(sel)
        if desc and len(desc.get_text()) > 80:
            job["Job Description"] = clean(desc.get_text(" "))[:4000]
            break
    log.debug(f"  Desc chars  → {len(job['Job Description'])}")

    # Experience / Qualifications from description
    job["Job Experience"]     = extract_experience(job["Job Description"])
    job["Job Qualifications"] = extract_qualifications(job["Job Description"])
    log.debug(f"  Experience  → {job['Job Experience']}")
    log.debug(f"  Quals       → {job['Job Qualifications'][:60] if job['Job Qualifications'] else ''}")

    # Salary fallback
    if not job["Salary Range"]:
        job["Salary Range"] = extract_salary(text)
    log.debug(f"  Salary      → {job['Salary Range']}")

    # Apply
    apply_a = sp.select_one("a[href*='apply'], a.btn-apply, a[class*='apply']")
    if apply_a:
        href = apply_a.get("href", "")
        job["Application"] = href if href.startswith("http") else base + href
    else:
        job["Application"] = url
    log.debug(f"  Application → {job['Application']}")

    job["Estimated Deadline"] = est_deadline(job["Date Posted"])
    return job


# ─────────────────────────────────────────────────────────────────
# SCRAPER 3 — CVPeopleAfrica.com
# ─────────────────────────────────────────────────────────────────
def scrape_cvpeople(browser):
    SOURCE = "CVPeopleAfrica.com"
    BASE   = "https://www.cvpeopleafrica.com"
    log.info(f"\n{'═'*60}")
    log.info(f"  SCRAPER 3 — {SOURCE}")
    log.info(f"{'═'*60}")
    jobs = []

    for pg in range(1, MAX_PAGES + 1):
        url = f"{BASE}/country/zambia" + (f"?page={pg}" if pg > 1 else "")
        log.info(f"\n  📄 Listing page {pg}/{MAX_PAGES}: {url}")
        page = browser.get(url)
        html = get_html(page)
        if not html:
            break

        sp    = BeautifulSoup(html, "lxml")
        cards = sp.select(
            "div.job-item, article.job, div[class*='job-card'], div[class*='vacancy']"
        )

        # Fallback: raw links
        if not cards:
            links = sp.select("a[href*='/job/'], a[href*='/vacancy/'], a[href*='/jobs/']")
            log.info(f"  Fallback: found {len(links)} raw links")
            for a in links[:30]:
                job = empty_job()
                job["Source Site"] = SOURCE
                job["Job Title"]   = clean(a.get_text())
                href               = a["href"]
                job["Job URL"]     = href if href.startswith("http") else BASE + href
                job["Job Location"]= "Zambia"
                job["Application"] = job["Job URL"]
                job["Estimated Deadline"] = est_deadline("")
                if job["Job Title"]:
                    log.info(f"  ┌ Job: {job['Job Title']}")
                    log_job(job)
                    jobs.append(job)
            if not links:
                break
            continue

        log.info(f"  Found {len(cards)} job card(s)")
        for i, card in enumerate(cards, 1):
            job = empty_job()
            job["Source Site"] = SOURCE

            sel_map = {
                "Job Title":   "h2, h3, .job-title",
                "Company Name":".company, .employer, .recruiter",
                "Job Location":".location, .city, .region",
                "Date Posted": ".date, time, .posted",
                "Salary Range":".salary, .remuneration",
                "Job Field":   ".category, .field, .sector",
            }
            for field, sel in sel_map.items():
                el = card.select_one(sel)
                if el:
                    job[field] = clean(el.get_text())

            a = card.select_one("a[href]")
            if a:
                href = a["href"]
                job["Job URL"] = href if href.startswith("http") else BASE + href

            desc = card.select_one("p, .excerpt, .summary, .job-excerpt")
            if desc:
                job["Job Description"] = clean(desc.get_text())[:2000]

            apply_el = card.select_one("a[href*='apply'], a.btn")
            if apply_el:
                href = apply_el.get("href", "")
                job["Application"] = href if href.startswith("http") else BASE + href
            elif job["Job URL"]:
                job["Application"] = job["Job URL"]

            job["Job Location"]       = job["Job Location"] or "Zambia"
            job["Date Posted"]        = normalise_date(job["Date Posted"])
            job["Estimated Deadline"] = est_deadline(job["Date Posted"])
            job["Job Experience"]     = extract_experience(job["Job Description"])
            job["Job Qualifications"] = extract_qualifications(job["Job Description"])

            if job["Job Title"]:
                log.info(f"  ┌ Job {i}: {job['Job Title']}")
                log_job(job)
                jobs.append(job)

    log.info(f"\n  ✅ {SOURCE} — total jobs scraped: {len(jobs)}")
    return jobs


# ─────────────────────────────────────────────────────────────────
# SCRAPER 4 — JobSearchZM.com
# ─────────────────────────────────────────────────────────────────
def scrape_jobsearchzm(browser):
    SOURCE = "JobSearchZM.com"
    BASE   = "https://jobsearchzm.com"
    log.info(f"\n{'═'*60}")
    log.info(f"  SCRAPER 4 — {SOURCE}")
    log.info(f"{'═'*60}")
    jobs = []

    for pg in range(1, MAX_PAGES + 1):
        url = BASE + (f"?page={pg}" if pg > 1 else "")
        log.info(f"\n  📄 Listing page {pg}/{MAX_PAGES}: {url}")
        page = browser.get(url, wait="domcontentloaded")
        html = get_html(page)
        if not html:
            break

        sp    = BeautifulSoup(html, "lxml")
        items = sp.select("article, .job-listing, .listing, div.job, li.job-item")
        if not items:
            items = sp.select("a[href*='/?p='], a[href*='/job/']")
        if not items:
            log.info("  No items found — stopping")
            break

        log.info(f"  Found {len(items)} item(s)")
        for i, item in enumerate(items, 1):
            job = empty_job()
            job["Source Site"] = SOURCE

            a = (item.select_one("a[href]")
                 if getattr(item, "name", "") != "a"
                 else item)
            if a:
                job["Job Title"] = clean(a.get_text())
                href = a.get("href", "")
                job["Job URL"] = href if href.startswith("http") else BASE + href

            sel_map = {
                "Company Name": ".company, .employer, small",
                "Job Location": ".location, .city",
                "Date Posted":  ".date, time, .posted-on",
                "Salary Range": ".salary",
                "Job Type":     ".job-type, .type",
                "Job Field":    ".category, .sector",
            }
            for field, sel in sel_map.items():
                el = item.select_one(sel)
                if el:
                    job[field] = clean(el.get_text())

            job["Job Location"]       = job["Job Location"] or "Zambia"
            job["Date Posted"]        = normalise_date(job["Date Posted"])
            job["Estimated Deadline"] = est_deadline(job["Date Posted"])
            job["Application"]        = job["Job URL"]

            if job["Job Title"]:
                log.info(f"  ┌ Job {i}: {job['Job Title']}")
                log_job(job)
                jobs.append(job)

    log.info(f"\n  ✅ {SOURCE} — total jobs scraped: {len(jobs)}")
    return jobs


# ─────────────────────────────────────────────────────────────────
# SCRAPER 5 — GreatZambiaJobs.com
# ─────────────────────────────────────────────────────────────────
def scrape_greatzambiajobs(browser):
    SOURCE = "GreatZambiaJobs.com"
    BASE   = "https://www.greatzambiajobs.com"
    log.info(f"\n{'═'*60}")
    log.info(f"  SCRAPER 5 — {SOURCE}")
    log.info(f"{'═'*60}")
    jobs = []

    for pg in range(1, MAX_PAGES + 1):
        url = f"{BASE}/jobs/" + (f"?paged={pg}" if pg > 1 else "")
        log.info(f"\n  📄 Listing page {pg}/{MAX_PAGES}: {url}")
        page = browser.get(url, wait="domcontentloaded")
        html = get_html(page)
        if not html:
            break

        sp    = BeautifulSoup(html, "lxml")
        cards = sp.select("article, .job-listing, div[class*='job']")
        if not cards:
            log.info("  No cards found — stopping")
            break

        log.info(f"  Found {len(cards)} card(s)")
        for i, card in enumerate(cards, 1):
            job = empty_job()
            job["Source Site"] = SOURCE

            a = card.select_one("h2 a, h3 a, .job-title a, a[rel='bookmark']")
            if a:
                job["Job Title"] = clean(a.get_text())
                href = a.get("href", "")
                job["Job URL"] = href if href.startswith("http") else BASE + href

            sel_map = {
                "Company Name": ".company, .employer, .job-company",
                "Job Location": ".location, .job-location, [class*='location']",
                "Date Posted":  ".date, time, .job-date, .posted",
                "Job Type":     ".job-type, .employment-type",
                "Salary Range": ".salary, .job-salary",
                "Job Field":    ".category, .job-category",
            }
            for field, sel in sel_map.items():
                el = card.select_one(sel)
                if el:
                    job[field] = clean(el.get_text())

            excerpt = card.select_one("p, .excerpt, .summary, .job-excerpt")
            if excerpt:
                job["Job Description"] = clean(excerpt.get_text())[:2000]

            logo = card.select_one("img[src*='logo'], img[class*='logo'], .company-logo img")
            if logo:
                job["Company Logo"] = logo.get("src", "")

            job["Job Location"]       = job["Job Location"] or "Zambia"
            job["Date Posted"]        = normalise_date(job["Date Posted"])
            job["Estimated Deadline"] = est_deadline(job["Date Posted"])
            job["Application"]        = job["Job URL"]
            job["Job Experience"]     = extract_experience(job["Job Description"])
            job["Job Qualifications"] = extract_qualifications(job["Job Description"])

            if job["Job Title"]:
                log.info(f"  ┌ Job {i}: {job['Job Title']}")
                log_job(job)
                jobs.append(job)

    log.info(f"\n  ✅ {SOURCE} — total jobs scraped: {len(jobs)}")
    return jobs


# ─────────────────────────────────────────────────────────────────
# WORDPRESS POST
# ─────────────────────────────────────────────────────────────────
def post_to_wordpress(job, wp_url, wp_user, wp_pass):
    endpoint = wp_url.rstrip("/") + "/wp-json/wp/v2/job_listing"
    data = {
        "title":   job.get("Job Title", ""),
        "content": job.get("Job Description", ""),
        "status":  "publish",
        "meta": {
            "_job_location":    job.get("Job Location", ""),
            "_company_name":    job.get("Company Name", ""),
            "_company_website": job.get("Company Website") or job.get("Company URL", ""),
            "_company_logo":    job.get("Company Logo", ""),
            "_job_salary":      job.get("Salary Range", ""),
            "_job_expires":     job.get("Deadline") or job.get("Estimated Deadline", ""),
            "_job_source_url":  job.get("Job URL", ""),
            "_application":     job.get("Application", ""),
        },
    }
    try:
        r = requests.post(
            endpoint, json=data,
            auth=(wp_user, wp_pass), timeout=20,
        )
        if r.status_code in (200, 201):
            log.info(f"  ✓ WP posted: {job['Job Title'][:60]}")
            return True
        else:
            log.warning(f"  ✗ WP error {r.status_code}: {r.text[:150]}")
            return False
    except Exception as e:
        log.error(f"  ✗ WP request failed: {e}")
        return False


# ─────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1A3C6E")
ALT_FILL  = PatternFill("solid", fgColor="EEF2FB")
WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)
THIN      = Side(style="thin", color="C8D3E8")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS = {
    "Job Title": 36, "Job Type": 14, "Job Qualifications": 24,
    "Job Experience": 20, "Job Location": 18, "Job Field": 22,
    "Date Posted": 13, "Deadline": 13, "Job Description": 65,
    "Application": 42, "Company URL": 42, "Company Name": 28,
    "Company Logo": 45, "Company Industry": 22, "Company Founded": 15,
    "Company Type": 15, "Company Website": 35, "Company Address": 30,
    "Company Details": 35, "Job URL": 45, "Estimated Deadline": 18,
    "Salary Range": 22, "Source Site": 22,
}


def write_excel(jobs, path):
    log.info(f"\n📊 Writing {len(jobs)} jobs to Excel → {path}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zambia Jobs"
    ws.freeze_panes = "A2"

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(col, 18)
    ws.row_dimensions[1].height = 30

    for ri, job in enumerate(jobs, 2):
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        for ci, col in enumerate(COLUMNS, 1):
            val = job.get(col, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = CELL_FONT
            c.fill      = fill
            c.border    = BORDER
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(col == "Job Description"),
            )

    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20
    ws2["A1"] = "Zambia Jobs Scraper v2.0 — Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="1A3C6E")
    ws2["A3"] = "Generated:"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A4"] = "Total Unique Jobs:"
    ws2["B4"] = len(jobs)

    sources = {}
    for j in jobs:
        s = j.get("Source Site", "Unknown")
        sources[s] = sources.get(s, 0) + 1

    ws2["A6"] = "Breakdown by Source"
    ws2["A6"].font = Font(bold=True)
    for i, (src, cnt) in enumerate(sources.items(), 7):
        ws2[f"A{i}"] = f"  • {src}"
        ws2[f"B{i}"] = cnt

    # ── Sites reference sheet
    ws3 = wb.create_sheet("Sites Reference")
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 42
    ws3.column_dimensions["C"].width = 34
    for ci, hdr in enumerate(["Source Site", "URL", "Notes"], 1):
        c = ws3.cell(row=1, column=ci, value=hdr)
        c.font = Font(bold=True, color="1A3C6E")

    sites = [
        ("GoZambiaJobs.com",    "https://gozambiajobs.com",         "Zambia's #1 — since 2011, highest volume"),
        ("ZambiaJob.com",       "https://www.zambiajob.com",         "AfricaWork network, 4000+ listings"),
        ("CVPeopleAfrica.com",  "https://www.cvpeopleafrica.com",    "Regional recruitment agency, verified listings"),
        ("JobSearchZM.com",     "https://jobsearchzm.com",           "Local aggregator, SME employers"),
        ("GreatZambiaJobs.com", "https://www.greatzambiajobs.com",   "Fast-updated regional board"),
    ]
    for ri, (name, url, note) in enumerate(sites, 2):
        ws3.cell(row=ri, column=1, value=name)
        ws3.cell(row=ri, column=2, value=url)
        ws3.cell(row=ri, column=3, value=note)

    wb.save(path)
    log.info(f"✅  Saved → {path}")


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Zambia Jobs Scraper v2.0 — MimusJobs")
    parser.add_argument("--max-pages", type=int, default=MAX_PAGES,
                        help=f"Pages to crawl per site (default {MAX_PAGES})")
    parser.add_argument("--output", default=OUTPUT_FILE,
                        help=f"Output Excel file (default {OUTPUT_FILE})")
    parser.add_argument("--post-wp", action="store_true",
                        help="Post scraped jobs to WordPress via REST API")
    parser.add_argument("--sites", nargs="+",
                        choices=["gozambia","zambiajob","cvpeople","jobsearchzm","greatzambia","all"],
                        default=["all"],
                        help="Which sites to scrape (default: all)")
    args = parser.parse_args()

    global MAX_PAGES
    MAX_PAGES = args.max_pages

    wp_url  = os.environ.get("WP_URL", "")
    wp_user = os.environ.get("WP_USER", "")
    wp_pass = os.environ.get("WP_PASS", "")

    run_all    = "all" in args.sites
    site_flags = {
        "gozambia":    run_all or "gozambia"    in args.sites,
        "zambiajob":   run_all or "zambiajob"   in args.sites,
        "cvpeople":    run_all or "cvpeople"     in args.sites,
        "jobsearchzm": run_all or "jobsearchzm" in args.sites,
        "greatzambia": run_all or "greatzambia" in args.sites,
    }

    log.info("═" * 60)
    log.info("  ZAMBIA JOBS SCRAPER v2.0 — MimusJobs")
    log.info(f"  Pages/site : {MAX_PAGES}")
    log.info(f"  Output     : {args.output}")
    log.info(f"  Post to WP : {'Yes — ' + wp_url if args.post_wp and wp_url else 'No'}")
    log.info(f"  Sites      : {[k for k,v in site_flags.items() if v]}")
    log.info(f"  Started    : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("═" * 60)

    all_jobs = []

    scrapers = [
        ("gozambia",    scrape_gozambiajobs),
        ("zambiajob",   scrape_zambiajob),
        ("cvpeople",    scrape_cvpeople),
        ("jobsearchzm", scrape_jobsearchzm),
        ("greatzambia", scrape_greatzambiajobs),
    ]

    with Browser() as browser:
        for key, scraper_fn in scrapers:
            if site_flags.get(key):
                try:
                    jobs = scraper_fn(browser)
                    all_jobs.extend(jobs)
                    log.info(f"  Running total: {len(all_jobs)} jobs so far")
                except Exception as e:
                    log.error(f"  Scraper '{key}' crashed: {e}", exc_info=True)

    # Deduplicate
    seen, unique = set(), []
    for j in all_jobs:
        key = j.get("Job URL") or j.get("Job Title", "")
        if key and key not in seen:
            seen.add(key)
            unique.append(j)

    dupes = len(all_jobs) - len(unique)
    log.info(f"\n📊 Total collected : {len(all_jobs)}")
    log.info(f"   Duplicates removed: {dupes}")
    log.info(f"   Unique jobs       : {len(unique)}")

    # WordPress posting
    if args.post_wp:
        if wp_url and wp_user and wp_pass:
            log.info(f"\n📤 Posting to WordPress: {wp_url}")
            ok = sum(
                post_to_wordpress(j, wp_url, wp_user, wp_pass)
                for j in unique
            )
            log.info(f"  Posted {ok}/{len(unique)} jobs successfully")
        else:
            log.warning("  --post-wp set but WP_URL / WP_USER / WP_PASS not all set in env")

    write_excel(unique, args.output)

    log.info(f"\n🏁 Done — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
